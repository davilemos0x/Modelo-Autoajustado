import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import psycopg2
import re
import statistics

now = datetime.now()

log = open('log.txt', 'a')
log.write("------ Excecução iniciada em: " + str(now) + "\n\n")


try:
#Conecta com o BD
	con = psycopg2.connect("host='localhost' port='5432' dbname='db_2019' user='postgres' password='12345'")
	cur = con.cursor()
	#cur.execute("SELECT * FROM potreiro")
	#records = cur.fetchall()
	#import pprint
	#pprint.pprint(records)
	# cur.execute("SELECT * FROM medicao")
	# con.commit()
	
	    
except:
	log.write("\nFalha na conexão com o BD!")

wb = load_workbook(filename='2020_planilhas_annoni.xlsx', data_only=True, read_only=False)
#wb = load_workbook(filename='Pastagem.xlsx', data_only=True)
planilha=['P 20 RECUP', 'P 21 RECUP']

for o in planilha:

	pattern= re.compile(r'\d+')
	num= re.findall(pattern, o)

	pattern= re.compile(r'I')
	desc = re.findall(pattern, o)
	
	if not desc:
		desc='R'
	else:
		desc='I'

	
	cur.execute("SELECT \"id\" from \"potreiro\" where \"nome\" like %s;", (o.lower()+'%',))
	
	try:
		idpotreiro = cur.fetchone()[0]
		con.commit()

	except:
		query =  "INSERT INTO potreiro (nome, numero, descricao) VALUES (%s, %s, %s) RETURNING id;"
		data = (o.lower(), num[0], desc)

		cur.execute(query, data)
		

		idpotreiro = cur.fetchone()[0]
		con.commit()
	
	p=wb[o]
	y = 0
	for x in range(9,99999999,16):
		d = p.cell(row=1, column=x)
		#print(d)
		if d.value != None:
			#print(d.strftime("%d/%m/%y"))
			print (d.value)

			dt=d.value

			if type(dt) is not datetime:
					dt = None
					log.write(o + " - Data inválida: " + str(d.value) + " Na célula : " + str(d) + "\n")
			
			cur.execute("SELECT \"id\" from \"medicao\" where \"data\"= %s;", (dt,))
			
			try:
				
				idmedicao = cur.fetchone()[0]
				con.commit()

			except:

				query =  "INSERT INTO medicao (data) VALUES (%s) RETURNING id;"
				data = (dt,)

				cur.execute(query, data)
				idmedicao = cur.fetchone()[0]
				con.commit()

			n = 6
			m = 14
			
			#if p.cell(row=6, column=x-7).value != None :
			
			print("Dentro da Gaiola - " + o)
			
			matriz1 = []
			
			tam=-1	
			for i in range(n):
			    valido=0
			    for j in range(m):
			        if isinstance(p.cell(row=i+6, column=j+2+y).value, int) or isinstance(p.cell(row=i+6, column=j+2+y).value, float):
			        	valido=1

			        else:
			        	p.cell(row=i+6, column=j+2+y).value=None

			    if valido==1:
			    	tam+=1
			    	matriz1.append([])

			    	for j in range(m):
			    		matriz1[tam].append(p.cell(row=i+6, column=j+2+y).value)

			    	matriz1[tam].append(i+1)
			

			
			for i in range(len(matriz1)):
			    for j in range(len(matriz1[i])):
			        print(matriz1[i][j], end=" ")
			    print ("\n")
		

			for i in range(len(matriz1)):
				try:
					mediana= statistics.median(matriz1[i][0:5])

				except:
					mediana=None

				idsubarea=None

				if idpotreiro==1 and len(matriz1)==3:
					idsubarea=25+i

				if idpotreiro==1 and len(matriz1)==6:
					idsubarea=1+i

				if idpotreiro==2 and len(matriz1)==3:
					idsubarea=28+i

				if idpotreiro==2 and len(matriz1)==6:
					idsubarea=7+i

				if idpotreiro==3 and len(matriz1)==3:
					idsubarea=31+i

				if idpotreiro==3 and len(matriz1)==6:
					idsubarea=13+i

				if idpotreiro==4 and len(matriz1)==3:
					idsubarea=34+i

				if idpotreiro==4 and len(matriz1)==6:
					idsubarea=19+i

				query =  "INSERT INTO pastagem (idmedicao, idpotreiro, dentrofora, gaiola, a1, a2, a3, a4, a5, media, mediana, pvtotal, pvsubamostra, psanoni, psoutras, psmorto, msanoni, msoutras, mstotal, ponto, idsubarea) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id;"
				data = (idmedicao , idpotreiro, 'D', matriz1[i][14], matriz1[i][0], matriz1[i][1], matriz1[i][2], matriz1[i][3], matriz1[i][4], matriz1[i][5], mediana, matriz1[i][6], matriz1[i][7], matriz1[i][8], matriz1[i][9], matriz1[i][10], matriz1[i][11], matriz1[i][12], matriz1[i][13], None, idsubarea )

				cur.execute(query, data)
				#idpastagem = cur.fetchone()[0]
				con.commit()

				#else:
					#print("NAO VALIDO------" + str(matriz1[i]) + "-------" + str(type(matriz1[i][0])))
					#print ("\n")
			

			#if p.cell(row=18, column=x-7).value != None :
				
			print("Fora da Gaiola - " + o)
			
			matriz2 = []
			
			tam=-1	
			for i in range(n):
				valido=0
				for j in range(m):
					if isinstance(p.cell(row=i+18, column=j+2+y).value, int) or isinstance(p.cell(row=i+18, column=j+2+y).value, float):
						valido=1
					else:
						p.cell(row=i+18, column=j+2+y).value=None

				if valido==1:
					tam+=1
					matriz2.append([])
					for j in range(m):
						matriz2[tam].append(p.cell(row=i+18, column=j+2+y).value)
					matriz2[tam].append(i+1)
		
			
			for i in range(len(matriz2)):
			    for j in range(len(matriz2[i])):
			        print(matriz2[i][j], end=" ")
			    print ("\n")
		

			for i in range(len(matriz2)):
				try:
					mediana= statistics.median(matriz2[i][0:5])

				except:
					mediana=None

				idsubarea=None

				if idpotreiro==1 and len(matriz2)==3:
					idsubarea=25+i

				if idpotreiro==1 and len(matriz2)==6:
					idsubarea=1+i

				if idpotreiro==2 and len(matriz2)==3:
					idsubarea=28+i

				if idpotreiro==2 and len(matriz2)==6:
					idsubarea=7+i

				if idpotreiro==3 and len(matriz2)==3:
					idsubarea=31+i

				if idpotreiro==3 and len(matriz2)==6:
					idsubarea=13+i

				if idpotreiro==4 and len(matriz2)==3:
					idsubarea=34+i

				if idpotreiro==4 and len(matriz2)==6:
					idsubarea=19+i

				query =  "INSERT INTO pastagem (idmedicao, idpotreiro, dentrofora, gaiola, a1, a2, a3, a4, a5, media, mediana, pvtotal, pvsubamostra, psanoni, psoutras, psmorto, msanoni, msoutras, mstotal, ponto, idsubarea) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id;"
				data = (idmedicao , idpotreiro, 'F', matriz2[i][14], matriz2[i][0], matriz2[i][1], matriz2[i][2], matriz2[i][3], matriz2[i][4], matriz2[i][5], mediana, matriz2[i][6], matriz2[i][7], matriz2[i][8], matriz2[i][9], matriz2[i][10], matriz2[i][11], matriz2[i][12], matriz2[i][13], None, idsubarea )

				cur.execute(query, data)
				#idpastagem = cur.fetchone()[0]
				con.commit()
				#else:
					#print("NAO VALIDO------" + str(matriz2[i]) + "-------" + str(type(matriz2[i][0])))
					#print ("\n")  
			
			matriz3 = []
			i=0
			salto=[0,3,5]
			while 1:
				matriz3.append([])
				for j in salto:
					matriz3[i].append(p.cell(row=i+38, column=j+2+y).value)
					#print(p.cell(row=i+38, column=j+2+y).value)
					#print(matriz3[i][j])

				if (matriz3[i][0] is None) and (matriz3[i][1] is None) and (matriz3[i][2] is None ):
					break
				else:
					if i>=2:
						query =  "INSERT INTO animais (idmedicao, idpotreiro, datapesagem, identificacao, pesoorigem, pesofinal, dataentrada, datasaida) VALUES (%s, %s, %s, %s, %s, %s, %s, %s) RETURNING id;"
						data = (idmedicao , idpotreiro, None, matriz3[i][0], matriz3[i][1], matriz3[i][2], matriz3[0][1], matriz3[0][2])

						cur.execute(query, data)
						con.commit()


				i+=1


			y += 16
		
		else:
			
			log.write(o + " - Parou ao ler: " + str(d.value) + " Na célula : " + str(d) + "\n")
			
			break



con.close()

now2 = datetime.now()
log.write("\n------ Excecução concluída em: " + str(now2))
log.write("\n------ Tempo de excecução: " + str(now2-now) + "\n\n\n")
log.close()
	


