import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date, datetime
import psycopg2
import re
import statistics
import xlrd

now = datetime.now()

log = open('log.txt', 'a')
log.write("------ Excecução iniciada em: " + str(now) + "\n\n")


try:
#Conecta com o BD
	con = psycopg2.connect("host='localhost' port='5432' dbname='db_2019' user='postgres' password='12345'")
	cur = con.cursor()	
		
except:
	log.write("\nFalha na conexão com o BD!")

#wb = load_workbook(filename='Planilha_Dados_de_2007_a_2016.xlsx', data_only=True, read_only=False)
book = xlrd.open_workbook("Dados_2007_2016_OK.xlsx")
sheet = book.sheet_by_name("Planilha1")

pk = 1
'''
#Inserir dados das medições, para isso tem que trocar a tabela de dados antes
resp = None
contador = 0
pk2 = 1
for r in range(1, sheet.nrows):
	id = pk
	data = xlrd.xldate.xldate_as_datetime(sheet.cell(r,2).value, book.datemode)
	responsavel = resp
	
	if (contador <= 8):
		id_medicao = pk2
	else:
		id_medicao = pk2 + 1
		pk2 = pk2 + 1
		contador = 0
	
	query = "INSERT INTO medicao_cristina (id, id_medicao, data, responsavel) VALUES (%s, %s, %s, %s)"
	values = (id, id_medicao, data, responsavel)

	cur.execute(query, values)
	con.commit()
	contador = contador + 1
	pk = pk + 1
	print(id, id_medicao, data, responsavel)
	print('')

 #Inserir dados dos potreiros
'''
'''
cobert = None
obs = None
geometria = None

for r in range(1, 10):
	id = pk
	potreiro = sheet.cell(r,1).value
	tratamento = sheet.cell(r,0).value
	cobertura = cobert
	observaçoes = obs
	poligono = geometria
	
	query = "INSERT INTO potreiro_cristina (id, potreiro, tratamento, cobertura, observaçoes, poligono) VALUES (%s, %s, %s, %s, %s, %s)"
	values = (id, potreiro, tratamento, cobertura, observaçoes, poligono)

	cur.execute(query, values)
	con.commit()
	pk = pk + 1 
	print(id, potreiro, tratamento, cobertura, observaçoes, poligono, end=" ")
'''
 #Inserir dados da pastagem

point = None
idsa = None
id = 1
idmedicao = 1
contador = 0
a1 = 0
a2 = 0
a3 = 0
a4 = 0
a5 = 0

for r in range(1, sheet.nrows):

	if sheet.cell(r,6).value == '':
		a1 = 0
	else:
		a1 = sheet.cell(r,6).value
	if sheet.cell(r,7).value == '':
		a2 = 0
	else:
		a2 = sheet.cell(r,7).value
	if sheet.cell(r,8).value == '':
		a3 = 0
	else:
		a3 = sheet.cell(r,8).value
	if sheet.cell(r,9).value == '':
		a4 = 0
	else:
		a4 = sheet.cell(r,9).value
	if sheet.cell(r,10).value == '':
		a5 = 0
	else:
		a5 = sheet.cell(r,10).value
	if sheet.cell(r,12).value == '':
		mstotal = 0
	else:
		mstotal = sheet.cell(r,12).value
	if sheet.cell(r,13).value == '':
		densid = 0
	else:
		densid = sheet.cell(r,13).value
	if sheet.cell(r,11).value == '':
		media = 0
	else:
		media = sheet.cell(r,11).value
		
	id = pk
	dados = [a1, a2, a3, a4, a5]
	#print("alturas", a1, a2, a3, a4, a5)
	mediana = statistics.median([a1, a2, a3, a4, a5])
	ponto = point
	idsubarea = idsa
	idpotreiro = sheet.cell(r,1).value
	dentrofora = sheet.cell(r,3).value
	gaiola = sheet.cell(r,2).value
	
	#data = xlrd.xldate.xldate_as_datetime(sheet.cell(r,2).value, book.datemode)

		
	ano = int(sheet.cell(r,5).value)
	mes = sheet.cell(r,4).value
	
	#print(mes, ano)
	if ano == 2007:
		if mes == "JUN":
			idmedicao = idmedicao
			print(sheet.cell(r,4).value, sheet.cell(r,5).value, id, idmedicao, idpotreiro, dentrofora, gaiola, a1, a2, a3, a4, a5, media, mediana, mstotal, densid, ponto, idsubarea)
			contador = 72
		else:
			if (contador <= 72):
				idmedicao = idmedicao
				print(sheet.cell(r,4).value, sheet.cell(r,5).value, id, idmedicao, idpotreiro, dentrofora, gaiola, a1, a2, a3, a4, a5, media, mediana, mstotal, densid, ponto, idsubarea)		
			else:
				idmedicao = idmedicao + 1
				print(sheet.cell(r,4).value, sheet.cell(r,5).value, id, idmedicao, idpotreiro, dentrofora, gaiola, a1, a2, a3, a4, a5, media, mediana, mstotal, densid, ponto, idsubarea)		
				contador = 1						
	elif ano == 2012:
		if mes == "AGO":
			idmedicao = 17
			print(sheet.cell(r,4).value, sheet.cell(r,5).value, id, idmedicao, idpotreiro, dentrofora, gaiola, a1, a2, a3, a4, a5, media, mediana, mstotal, densid, ponto, idsubarea)
			contador = 72
		else:
			if (contador <= 72):
				idmedicao = idmedicao
				print(sheet.cell(r,4).value, sheet.cell(r,5).value, id, idmedicao, idpotreiro, dentrofora, gaiola, a1, a2, a3, a4, a5, media, mediana, mstotal, densid, ponto, idsubarea)		
			else:
				idmedicao = idmedicao + 1
				print(sheet.cell(r,4).value, sheet.cell(r,5).value, id, idmedicao, idpotreiro, dentrofora, gaiola, a1, a2, a3, a4, a5, media, mediana, mstotal, densid, ponto, idsubarea)		
				contador = 1					
	else:
		if (contador <= 72):
			idmedicao = idmedicao
			print(sheet.cell(r,4).value, sheet.cell(r,5).value, id, idmedicao, idpotreiro, dentrofora, gaiola, a1, a2, a3, a4, a5, media, mediana, mstotal, densid, ponto, idsubarea)		
		else:
			idmedicao = idmedicao + 1
			print(sheet.cell(r,4).value, sheet.cell(r,5).value, id, idmedicao, idpotreiro, dentrofora, gaiola, a1, a2, a3, a4, a5, media, mediana, mstotal, densid, ponto, idsubarea)		
			contador = 1
			
	query = "INSERT INTO pastagem_cristina (id, idmedicao, idpotreiro, dentrofora, gaiola, a1, a2, a3, a4, a5, media, mediana, mstotal, densidade, ponto, idsubarea) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
	values = (id, idmedicao, idpotreiro, dentrofora, gaiola, a1, a2, a3, a4, a5, media, mediana, mstotal, densid, ponto, idsubarea)

	cur.execute(query, values)
	con.commit()
	#print('Salvo no banco')
		
	contador = contador + 1
	#print(contador)
	pk = pk + 1	  
	id = id + 1

con.close()

now2 = datetime.now()
log.write("\n------ Excecução concluída em: " + str(now2))
log.write("\n------ Tempo de excecução: " + str(now2-now) + "\n\n\n")
log.close()

